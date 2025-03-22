from .formatter import Formatter
from .utilities import shrink_path

import openpyxl as op


class Spreadsheet:
    def __init__(self, filename, tbb):
        self._tbb = tbb
        path = shrink_path(filename)
        self._filename = 'spreadsheets' / path.with_suffix('.xlsx')
        self._format_dirname = 'format' / path.with_suffix('')
        if not self._format_dirname.exists():
            self._format_dirname.mkdir(parents=True, exist_ok=True)
        assert(self._format_dirname.exists())


    def load_spreadsheet(self):
        print('Loading', self._filename)
        spreadsheet = op.load_workbook(self._filename, read_only=False)

        for table in self._tbb.tables:
            format_filename = self._format_dirname / f'table_{table.index}.json'
            reader = Formatter(format_filename)
            sheetname = reader.sheet_name
            sheet = spreadsheet[sheetname]
            
            data = bytearray()

            if sheet.max_row > 2:
                rows = iter(sheet)
                next(rows) # skip column headers 

                row = next(rows) # include address headers
                headers = []
                for cell in row[2:]:
                    headers.append(int(cell.value, 0x10))

                for row in rows:
                    col = 0
                    for cell in row[2:]:
                        data += reader.get(cell.value, headers[col])
                        col += 1

            table.build(data)

        spreadsheet.close()


    def dump_spreadsheet(self):
        print('Building', self._filename)
        spreadsheet = op.Workbook()
        del spreadsheet['Sheet']

        for table in self._tbb.tables:
            format_filename = self._format_dirname / f'table_{table.index}.json'
            reader = Formatter(format_filename)
            sheetname = reader.sheet_name
            
            sheet = spreadsheet.create_sheet(sheetname)
            def readline():
                addr0 = table.data.tell()
                line = [hex(addr0), entry]
                offsets = []
                row_offset = 0
                while table.data.tell() < addr0 + table.row_size:
                    row_offset = table.data.tell() - addr0
                    offsets.append(hex(row_offset))
                    line.append(reader.read(table.data, row_offset))
                return line, offsets

            entry = 0
            row = 1
            widths = None
            for _ in range(table.num_rows):
                addr0 = table.data.tell()
                line, offsets = readline()
                if row == 1:
                    widths = [4] * (len(offsets) + 2)
                    sheet.cell(row, 1, 'Addr')
                    sheet.cell(row, 2, 'Row#')
                    col = 3
                    header = [''] * len(offsets)
                    for i in range(len(header)):
                        v = reader.header(offsets[i])
                        sheet.cell(row, col, v)
                        widths[col-1] = max(widths[col-1], len(v))
                        col += 1
                    row += 1
                    col = 3
                    for value in offsets:
                        sheet.cell(row, col, value)
                        widths[col-1] = max(widths[col-1], len(value))
                        col += 1
                    row += 1

                for col, value in enumerate(line):
                    sheet.cell(row, col+1, value)
                    w = 4
                    for line in str(value).split('\n'):
                        wline = len(line)
                        if line and line[0].encode('utf8')[0] & 0x80:
                            wline += len(line) # shift-jis vary a lot in width
                        w = max(w, wline)
                    widths[col] = max(widths[col], w)

                addr1 = table.data.tell()
                assert(addr1 - addr0 == table.row_size)

                row += 1
                entry += 1

                for i, w in enumerate(widths):
                    col_letter = sheet.cell(1, i+1).column_letter
                    sheet.column_dimensions[col_letter].width = int(w + 2)

                sheet.freeze_panes = sheet.cell(3, 3)

        print('Writing spreadsheet', self._filename)
        if not self._filename.parent.exists():
            self._filename.parent.mkdir(parents=True)
        spreadsheet.save(self._filename)
